import os
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from pathlib import Path

# 設定範本路徑
BASE_DIR = Path(__file__).resolve().parent.parent.parent
TEMPLATE_PATH = BASE_DIR / "templates" / "worklog_template.xlsx"

class ExcelExporter:
    def __init__(self, target_filepath: str):
        """
        初始化匯出器
        :param target_filepath: 使用者選擇的匯出目標路徑 (例如: D:/每月報表/2026_工作日誌.xlsx)
        """
        self.target_filepath = target_filepath
        
        # 判斷目標檔案是否存在，若存在則載入追加，若不存在則載入範本另存
        if os.path.exists(self.target_filepath):
            self.wb = openpyxl.load_workbook(self.target_filepath)
        else:
            if not os.path.exists(TEMPLATE_PATH):
                raise FileNotFoundError(f"找不到範本檔案：{TEMPLATE_PATH}")
            self.wb = openpyxl.load_workbook(TEMPLATE_PATH)

    def _copy_sheet_format(self, source_sheet, target_sheet):
        """深層複製：將來源 Sheet 的欄寬、列高完全複製到目標 Sheet"""
        # 複製欄寬
        for col in source_sheet.column_dimensions:
            target_sheet.column_dimensions[col].width = source_sheet.column_dimensions[col].width
            
        # 複製列高
        for row in source_sheet.row_dimensions:
            target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height

    def append_monthly_data(self, month_str: str, data_rows: list):
        """
        建立當月新分頁並寫入資料
        :param month_str: 新分頁名稱 (例如: '工作報告11508')
        :param data_rows: 要寫入的資料列表 (List of Lists 或 Dicts)
        """
        # 1. 取得範本 Sheet (假設範本的 Sheet 名稱為 'Template' 或直接抓第一個)
        template_sheet = self.wb.worksheets[0]
        
        # 2. 如果該月份的 Sheet 已經存在，先刪除避免重疊 (或者你可以選擇拋出錯誤)
        if month_str in self.wb.sheetnames:
            del self.wb[month_str]

        # 3. 複製範本為新的 Sheet
        new_sheet = self.wb.copy_worksheet(template_sheet)
        new_sheet.title = month_str
        
        # 4. 確保欄寬完美複製 (copy_worksheet 有時不會拷貝欄寬)
        self._copy_sheet_format(template_sheet, new_sheet)

        # 5. 尋找開始寫入資料的那一列 (假設範本第1行為大標題，第2行為欄位名稱，我們從第3行開始寫)
        start_row = 3
        
        # 6. 寫入資料
        for row_idx, row_data in enumerate(data_rows, start=start_row):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = new_sheet.cell(row=row_idx, column=col_idx)
                cell.value = cell_value
                
                # 複製範本第一筆資料列 (start_row) 的樣式 (字體、框線、對齊)
                template_cell = template_sheet.cell(row=start_row, column=col_idx)
                if template_cell.has_style:
                    cell.font = copy(template_cell.font)
                    cell.border = copy(template_cell.border)
                    cell.fill = copy(template_cell.fill)
                    cell.number_format = copy(template_cell.number_format)
                    cell.alignment = copy(template_cell.alignment)

        # 7. 如果這是一個全新的檔案，把預設的 'Template' 隱藏或刪除，保持乾淨
        if len(self.wb.sheetnames) > 1 and template_sheet.title == 'Template':
            template_sheet.sheet_state = 'hidden'

    def save(self):
        """儲存 Excel 檔案"""
        self.wb.save(self.target_filepath)
        print(f"✅ Excel 檔案已成功儲存至：{self.target_filepath}")